#!/usr/bin/env python3
"""Keep the KXHMONTH decision workbook in sync with the model, in place.

    python3 xlsx_sync.py                      # refresh the current month's inputs
    python3 xlsx_sync.py --month 2026-08      # pin a month
    python3 xlsx_sync.py --build              # create the month sheet (layout + formulas)
    python3 xlsx_sync.py --charts             # rebuild that sheet's charts

Three modes on purpose. The daily refresh writes ONLY input cells — never a
formula, never formatting, never a chart — so anything you restyle in Excel
survives. --charts is opt-in for exactly that reason.

Run it after daily_check.py, which is what refreshes era5_daily.csv,
forecast_log/, data.js and the ensemble archive that this reads.
"""
import argparse, calendar, json, math, os, subprocess, sys
import datetime as dt
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference, Series

HERE = os.path.dirname(os.path.abspath(__file__))
BOOK = os.path.join(HERE, 'KXHMONTH_2026 (decision).xlsx')
KALSHI = "https://api.elections.kalshi.com/trade-api/v2"
FITLO = 1990

# ---- layout constants: these mirror the workbook you built; changing one
# ---- without changing the sheet will silently write into the wrong cells.
HDR = 17            # daily-log header row
PRE0 = 18           # first pre-month row (4 of them: market opens early)
M0 = 22             # first in-month row
COL = dict(date='A', actual='B', exp='C', dev='D', dev3='E', mtd='F',
           an1='G', an2='H', need='I', pyes='J', fairno='K', mkt='L',
           edge='M', vol='N', oi='O', signal='P', note='Q',
           pace='R', modelc='S')      # R,S are chart helpers
NAVY, BAND, GREEN, YELLOW, GREY = 'FF1F3864', 'FFD9E2F3', 'FFE2EFDA', 'FFFFF2CC', 'FFF2F2F2'
BLUE_TXT, WHITE, RED = 'FF0000FF', 'FFFFFFFF', 'FFC00000'
EXP_RED = 'FFE6B9B8'   # accent2 @ tint 0.6 — Expectation ran UNDER the actual
ANOM, CENT, PCT, NUM0, MONEY = '+0.000;-0.000;0.000', '0.0', '0.0%', '#,##0', '$#,##0.00'
F = lambda **kw: Font(name='Arial', **kw)
THIN = Side(style='thin', color='FFBFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLOCKS = [('A', 'C', 'D'), ('E', 'G', 'H'), ('I', 'K', 'L'), ('M', 'P', 'Q')]


def geom(nd):
    m1 = M0 + nd - 1
    cal = m1 + 3                       # m1+1 MEAN row, m1+2 blank, then banner
    return dict(m1=m1, cal=cal, h0=cal + 2, h1=cal + 1 + (2026 - FITLO),
                s0=cal + 2, s1=cal + 2 + nd)


def exp_cf(ws, nd):
    """Expectation (C) goes light red when the forecast came in UNDER the actual;
    it keeps its green fill when it met or beat it. Conditional formatting, so it
    re-evaluates itself as each day lands -- refresh() never has to touch styling.
    Idempotent: only ever added once per sheet."""
    rng = f'C{M0}:C{geom(nd)["m1"]}'
    if any(str(cf.sqref) == rng for cf in ws.conditional_formatting):
        return False
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND($B{M0}<>"",$C{M0}<>"",$C{M0}<$B{M0})'],
        fill=PatternFill('solid', bgColor=EXP_RED)))
    return True


# ------------------------------------------------------------------ sources
def era5():
    daily, clim, obs = defaultdict(dict), {}, {}
    for line in open(os.path.join(HERE, 'era5_daily.csv')):
        if line.startswith('#'):
            continue
        p = line.strip().split(',')
        if len(p) >= 4 and p[0][:4].isdigit():
            y, m, d = map(int, p[0].split('-'))
            daily[(y, m)][d] = float(p[3])
            clim[p[0][5:]] = float(p[2])
            obs[p[0]] = float(p[3])
    return daily, clim, obs


def ols(pairs):
    n = len(pairs); sx = sum(x for x, _ in pairs); sy = sum(y for _, y in pairs)
    sxx = sum(x * x for x, _ in pairs); sxy = sum(x * y for x, y in pairs)
    b = (n * sxy - sx * sy) / (n * sxx - sx * sx); a = (sy - b * sx) / n
    return a, b, math.sqrt(sum((y - a - b * x) ** 2 for x, y in pairs) / (n - 2)), n


def ols2(X, Y):
    n = len(Y); x1 = [x[0] for x in X]; x2 = [x[1] for x in X]
    A = [[n, sum(x1), sum(x2)],
         [sum(x1), sum(a * a for a in x1), sum(a * b for a, b in zip(x1, x2))],
         [sum(x2), sum(a * b for a, b in zip(x1, x2)), sum(a * a for a in x2)]]
    v = [sum(Y), sum(a * y for a, y in zip(x1, Y)), sum(a * y for a, y in zip(x2, Y))]
    for i in range(3):
        for j in range(i + 1, 3):
            f = A[j][i] / A[i][i]
            for q in range(3):
                A[j][q] -= f * A[i][q]
            v[j] -= f * v[i]
    c = [0.0] * 3
    for i in (2, 1, 0):
        c[i] = (v[i] - sum(A[i][q] * c[q] for q in range(i + 1, 3))) / A[i][i]
    return c, math.sqrt(sum((y - c[0] - c[1] * a - c[2] * b) ** 2
                            for (a, b), y in zip(X, Y)) / (n - 3))


def noaa(m):
    return {int(k): v['departure'] for k, v in
            json.load(open(os.path.join(HERE, f'noaa_m{m}.json')))['data'].items()}


def gistemp(m):
    g = {}
    for line in open(os.path.join(HERE, 'gistemp.txt')):
        p = line.split()
        if p and p[0].isdigit() and len(p) > 12 and p[m].replace('-', '').isdigit():
            g[int(p[0])] = int(p[m]) / 100.0
    return g


def forecasts(year, month, clim, obs):
    """Leakage-free: latest pull strictly before each day, anchored only on days
    that were published at pull time (pull_date - 2, ERA5's lag)."""
    logdir = os.path.join(HERE, 'forecast_log')
    pulls = []
    for f in sorted(os.listdir(logdir)):
        if f.startswith('fcst_'):
            s = f[5:13]
            pulls.append((f'{s[:4]}-{s[4:6]}-{s[6:]}', os.path.join(logdir, f)))
    out = {}
    for d in range(1, calendar.monthrange(year, month)[1] + 1):
        t = f'{year}-{month:02d}-{d:02d}'
        prior = [p for p in pulls if p[0] < t]
        if not prior:
            continue
        iso, path = prior[-1]
        gm = json.load(open(path))['global_mean_C']
        raw = {k: gm[k] - clim[k[5:]] for k in gm if k[5:] in clim}
        cut = (dt.date.fromisoformat(iso) - dt.timedelta(days=2)).isoformat()
        ov = [(obs[k], raw[k]) for k in raw if k in obs and k <= cut]
        if not ov or t not in raw:
            continue
        off = sum(o - r for o, r in ov) / len(ov)
        assert abs(off) < 0.5, f'anchor {off:+.3f} insane in {path}'
        out[d] = round(raw[t] + off, 4)
    return out


def curl_json(url):
    r = subprocess.run(['curl', '-sL', '--max-time', '60', '-A', 'Mozilla/5.0', url],
                       capture_output=True)
    return json.loads(r.stdout) if r.stdout else {}


def market(ticker):
    """Cached candles + a live top-up, so today's row is filled before the
    candle for today closes."""
    out = {}
    p = os.path.join(HERE, 'kalshi_data', f'candles_{ticker}.json')
    if os.path.exists(p):
        for c in json.load(open(p))['candlesticks']:
            t = dt.datetime.fromtimestamp(c['end_period_ts'], tz=dt.timezone.utc).strftime('%Y-%m-%d')
            px = c['price'].get('close_dollars')
            out[t] = {'close': round(float(px) * 100, 1) if px else None,
                      'vol': round(float(c.get('volume_fp') or 0), 2),
                      'oi': round(float(c.get('open_interest_fp') or 0), 2)}
    try:
        cd = curl_json(f"{KALSHI}/series/KXHMONTH/markets/{ticker}/candlesticks"
                       f"?start_ts=1743465600&end_ts={int(dt.datetime.now().timestamp())+86400}"
                       f"&period_interval=1440")
        for c in cd.get('candlesticks', []):
            t = dt.datetime.fromtimestamp(c['end_period_ts'], tz=dt.timezone.utc).strftime('%Y-%m-%d')
            px = c['price'].get('close_dollars')
            if px:
                out[t] = {'close': round(float(px) * 100, 1),
                          'vol': round(float(c.get('volume_fp') or 0), 2),
                          'oi': round(float(c.get('open_interest_fp') or 0), 2)}
    except Exception as e:
        print(f'  kalshi candles unavailable ({e}) — using cache')
    return out


def latest_ens(year, month):
    cal = os.path.join(HERE, 'July Calibration')
    best = None
    for f in sorted(os.listdir(cal)):
        if not f.startswith('ens_spread_'):
            continue
        try:
            e = json.load(open(os.path.join(cal, f)))
        except Exception:
            continue
        mem = e.get('member_month_mean_era5')
        # strict month/year match: a stale file would price the wrong market
        if not mem or e.get('month') != month or e.get('year') != year:
            continue
        best = {'run': e['pulled_at'][:10], 'n': e['n_members'],
                'cold': round(min(mem), 4), 'avg': round(sum(mem) / len(mem), 4),
                'hot': round(max(mem), 4),
                'breach': f"{e['noaa']['members_central_breach']}/{e['n_members']}",
                'p': round(e['noaa']['p_yes_pct'] / 100, 4)}
    return best


def pipeline_p(year, month):
    try:
        dj = json.loads(open(os.path.join(HERE, 'data.js')).read().split('=', 1)[1].rstrip(';\n'))
        M = dj['model']
        if M['year'] != year or M['cur'] != calendar.month_abbr[month].lower():
            return None
        cm = M[M['cur']]
        if cm.get('ens51'):
            return round(cm['ens51']['p_yes_pct'] / 100, 4)
        ens = cm.get('ens')
        if ens:
            mu = cm['a'] + cm['b'] * ens['mu_era5']
            sig = math.hypot(cm['b'] * ens['sd_f'], cm['sd'])
            thr = round(cm['record'] + 0.005, 3)
            return round(0.5 * math.erfc((thr - mu) / sig / math.sqrt(2)), 4)
    except Exception as e:
        print('  data.js P(YES) unavailable:', e)
    return None


def collect(year, month):
    daily, clim, obs = era5()
    nd = calendar.monthrange(year, month)[1]
    py, pm = (year - 1, 12) if month == 1 else (year, month - 1)
    mm = lambda y, m: sum(daily[(y, m)].values()) / len(daily[(y, m)])
    ns = noaa(month)
    a, b, sd, n = ols([(mm(y, month), ns[y]) for y in range(FITLO, year)])
    rec = max(v for y, v in ns.items() if y < year)
    gs = gistemp(month)
    ga, gb, gsd, _ = ols([(mm(y, month), gs[y]) for y in range(FITLO, year) if y in gs])
    tops = sorted((k for k in ns if k < year), key=lambda k: -ns[k])[:2]
    tops = sorted(tops, reverse=True)
    # sigma_f(k): the variance-collapse schedule
    coll = {}
    pmof = lambda y: mm(y - 1, 12) if month == 1 else mm(y, month - 1)
    Y = [mm(y, month) for y in range(FITLO, year)]
    for k in range(nd + 1):
        if k == 0:
            coll[k] = round(ols([(pmof(y), mm(y, month)) for y in range(FITLO, year)])[2], 4)
        elif k >= nd:
            coll[k] = 0.0
        else:
            X = [(pmof(y), sum(daily[(y, month)][d] for d in range(1, k + 1)) / k)
                 for y in range(FITLO, year)]
            coll[k] = round(ols2(X, Y)[1], 4)
    bias = []
    for m in range(1, 13):
        if (year, m) not in daily or len(daily[(year, m)]) < calendar.monthrange(year, m)[1]:
            continue
        s = noaa(m)
        if year not in s:
            continue
        am, bm, _, _ = ols([(mm(y, m), s[y]) for y in range(FITLO, year)])
        bias.append([calendar.month_abbr[m], round(mm(year, m), 4), s[year],
                     round(am + bm * mm(year, m), 4),
                     round(s[year] - (am + bm * mm(year, m)), 4)])
    # the assumed-rate default is the previous month's mean, but a sheet built
    # before that month has any ERA5 days would divide by zero -- walk back to
    # the most recent month that actually has data.
    prev_mean, yy, mmn = None, py, pm
    for _ in range(12):
        if daily.get((yy, mmn)):
            prev_mean = mm(yy, mmn)
            break
        yy, mmn = (yy - 1, 12) if mmn == 1 else (yy, mmn - 1)
    assert prev_mean is not None, 'no ERA5 month found for the assumed-rate default'

    tk = f'KXHMONTH-{year%100:02d}{calendar.month_abbr[month].upper()}'
    mk = market(tk)
    fc = forecasts(year, month, clim, obs)
    rows = []
    cur = dt.date(year, month, 1) - dt.timedelta(days=4)
    end = dt.date(year, month, nd)
    while cur <= end:
        iso = cur.isoformat()
        inm = (cur.year, cur.month) == (year, month)
        i = (cur - (dt.date(year, month, 1) - dt.timedelta(days=4))).days
        an = []
        for yy in tops:
            if inm:
                dsr = daily.get((yy, month), {})
                an.append(round(sum(dsr[q] for q in range(1, cur.day + 1)) / cur.day, 4)
                          if len(dsr) >= cur.day else None)
            else:
                an.append(daily.get((yy, cur.month), {}).get(cur.day))
        rows.append({'label': cur.strftime('%b %d'), 'day': cur.day if inm else None,
                     'actual': obs.get(iso), 'fcst': fc.get(cur.day) if inm else None,
                     'mkt': mk.get(iso, {}), 'an': an})
        cur += dt.timedelta(days=1)
    return dict(year=year, month=month, nd=nd, ticker=tk, tops=tops,
                mname=calendar.month_name[month], a=round(a, 4), b=round(b, 4),
                sd=round(sd, 4), n=n, rec=rec, thr=round(rec + 0.005, 3),
                bar=round((round(rec + 0.005, 3) - a) / b, 5),
                gis_rec=max(v for y, v in gs.items() if y < year),
                gis_thr=round(max(v for y, v in gs.items() if y < year) + 0.005, 3),
                prev_mean=round(prev_mean, 4), printed=ns.get(year),
                hist=[[y, round(mm(y, month), 4), ns[y]] for y in range(FITLO, year)],
                coll=coll, bias=bias, rows=rows, ens=latest_ens(year, month),
                pipe_p=pipeline_p(year, month),
                settle={6: 'Jul 9 2026 10:00 ET', 7: 'Aug 8 2026 10:00 ET',
                        8: 'Sep 8 2026 10:00 ET', 9: 'Oct 8 2026 10:00 ET'}.get(month, 'TBC'))


# ------------------------------------------------------------------ writing
def banner(ws, rng, text, size=10):
    ws.merge_cells(rng)
    c = ws[rng.split(':')[0]]
    c.value, c.font = text, F(size=size, bold=True, color=WHITE)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')


def kv(ws, blk, row, label, value, fmt=None, inp=False, key=False, bold=False):
    la, lb, vc = BLOCKS[blk]
    ws.merge_cells(f'{la}{row}:{lb}{row}')
    lc = ws[f'{la}{row}']
    lc.value, lc.font = label, F(size=9, bold=bold)
    lc.fill = PatternFill('solid', fgColor=BAND if bold else GREY)
    lc.border = BOX
    lc.alignment = Alignment(horizontal='left', indent=1)
    v = ws[f'{vc}{row}']
    v.value = value
    v.font = F(size=10, bold=True, color=BLUE_TXT if inp else 'FF000000')
    v.fill = PatternFill('solid', fgColor=YELLOW if key else (GREEN if inp else WHITE))
    v.border = BOX
    v.alignment = Alignment(horizontal='center')
    if fmt:
        v.number_format = fmt


def build_sheet(wb, d):
    g = geom(d['nd'])
    M1, CAL, H0, H1, S0, S1 = g['m1'], g['cal'], g['h0'], g['h1'], g['s0'], g['s1']
    name = f"{calendar.month_abbr[d['month']]}_{d['year']}"
    ws = wb.create_sheet(name)
    for L, w in zip('ABCDEFGHIJKLMNOPQRST',
                    [6.0, 9.5, 10.8, 7.7, 7.0, 8.5, 6.5, 8.2, 7.2, 7.0, 7.2, 10.2,
                     7.0, 6.7, 7.7, 5.7, 7.7, 6.5, 8.0, 8.8]):
        ws.column_dimensions[L].width = w

    ws.merge_cells('A1:Q1')
    t = ws['A1']
    t.value = f"{d['ticker']}  —  “Will {d['mname']} {d['year']} be the hottest {d['mname']} ever?”"
    t.font = F(size=14, bold=True, color=WHITE)
    t.fill = PatternFill('solid', fgColor=NAVY)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    ws.merge_cells('A2:Q2')
    st = ws['A2']
    st.value = (("SETTLED — NOAA printed %.2f" % d['printed']) if d['printed'] is not None
                else "LIVE") + \
        f"   ·   record {d['rec']:.2f}, needs a {d['rec']+0.01:.2f} print" \
        f"   ·   ERA5 bar {d['bar']:+.5f}/day   ·   fit 1990–{d['year']-1}, n={d['n']}" \
        f"   ·   settles {d['settle']}"
    st.font = F(size=9, bold=True, color=NAVY)
    st.fill = PatternFill('solid', fgColor=BAND)
    st.alignment = Alignment(horizontal='center')

    for i, nm in enumerate(['CONTRACT & TRANSLATION', 'CURRENT STATE', 'ENSEMBLE',
                            'MARKET & P&L']):
        la, _, vc = BLOCKS[i]
        banner(ws, f'{la}3:{vc}3', nm)

    HX, HY = f'$B${H0}:$B${H1}', f'$C${H0}:$C${H1}'
    kv(ws, 0, 4, 'Record (printed)', d['rec'], '0.00')
    kv(ws, 0, 5, 'Latent threshold', d['thr'], '0.000')
    kv(ws, 0, 6, 'a  (intercept)', d['a'], '0.0000')
    kv(ws, 0, 7, 'b  (slope)', d['b'], '0.0000')
    kv(ws, 0, 8, 'σ translation', d['sd'], '0.0000')
    kv(ws, 0, 9, 'ERA5 bar (full month)', '=($D$5-$D$6)/$D$7', '0.00000')
    kv(ws, 0, 10, 'σ_f  at k',
       f'=IFERROR(INDEX($F${S0}:$F${S1},MATCH($H$4,$E${S0}:$E${S1},0)),$D$8)', '0.0000')
    kv(ws, 0, 11, 'σ total', '=SQRT(($D$7*$D$10)^2+$D$8^2)', '0.0000')
    kv(ws, 0, 12, 'Fit check',
       f'=IF(AND(ABS(SLOPE({HY},{HX})-$D$7)<0.0005,ABS(INTERCEPT({HY},{HX})-$D$6)<0.0005,'
       f'ABS(STEYX({HY},{HX})-$D$8)<0.0005),"OK","DRIFT")')
    kv(ws, 0, 13, 'GISTEMP record', d['gis_rec'], '0.00')
    kv(ws, 0, 14, 'GISTEMP threshold', d['gis_thr'], '0.000')
    kv(ws, 0, 15, 'ONI (CPC)', 1.0, '0.0', inp=True)

    nd = d['nd']
    kv(ws, 1, 4, 'k  (days observed)', f'=COUNT($B${M0}:$B${M1})', '0')
    kv(ws, 1, 5, 'MTD mean', f'=IF($H$4=0,"",AVERAGE($B${M0}:$B${M1}))', ANOM)
    kv(ws, 1, 6, 'Needed /day over rest',
       f'=IF(OR($H$4=0,$H$4>={nd}),"",($D$9*{nd}-SUM($B${M0}:$B${M1}))/({nd}-$H$4))', ANOM)
    kv(ws, 1, 7, 'Assumed rate, remaining days', d['prev_mean'], ANOM, key=True)
    kv(ws, 1, 8, 'Projected full-month ERA5',
       f'=(SUM($B${M0}:$B${M1})+({nd}-$H$4)*$H$7)/{nd}', ANOM)
    kv(ws, 1, 9, 'Projected NOAA μ', '=$D$6+$D$7*$H$8', '0.000')
    kv(ws, 1, 10, 'P(YES) — sheet', '=1-NORMDIST($D$5,$H$9,$D$11,TRUE)', PCT)
    kv(ws, 1, 11, 'P(YES) — pipeline', d['pipe_p'], PCT, inp=True)
    kv(ws, 1, 12, 'P(YES) used', '=IF($H$11="",$H$10,$H$11)', PCT, bold=True)
    kv(ws, 1, 13, 'Fair NO ¢', '=100*(1-$H$12)', CENT, bold=True)
    kv(ws, 1, 14, 'Edge on NO ¢',
       f'=IF(COUNT($L${PRE0}:$L${M1})=0,"",LOOKUP(9.99E+307,$L${PRE0}:$L${M1})-100*$H$12)',
       CENT, bold=True)
    kv(ws, 1, 15, 'Verdict',
       '=IF($H$14="","",IF($H$14>=3,"NO CHEAP",IF($H$14<=-3,"NO RICH","FAIR")))', bold=True)

    e = d['ens']
    kv(ws, 2, 4, 'Run', e['run'] if e else '—')
    kv(ws, 2, 5, 'Members', e['n'] if e else None, '0')
    kv(ws, 2, 6, 'Coldest member', e['cold'] if e else None, ANOM)
    kv(ws, 2, 7, 'Average member', e['avg'] if e else None, ANOM)
    kv(ws, 2, 8, 'Hottest member', e['hot'] if e else None, ANOM)
    kv(ws, 2, 9, 'Members breaching record', e['breach'] if e else None)
    kv(ws, 2, 10, 'Ensemble P(YES)', e['p'] if e else None, PCT, inp=True)
    kv(ws, 2, 11, 'dev3 (latest)', f'=IFERROR(LOOKUP(9.99E+307,$E${PRE0}:$E${M1}),"")', ANOM)
    kv(ws, 2, 12, 'Spike signal',
       '=IF($L$11="","",IF($L$11>0.05,"WARM · YES↑",IF($L$11<-0.05,"COLD · YES↓","quiet")))')
    r = [x[4] for x in d['bias']]
    kv(ws, 2, 13, 'Bias — 2026 mean residual',
       round(sum(r) / len(r), 4) if r else None, '+0.0000;-0.0000')
    kv(ws, 2, 14, 'Apply bias to μ?', 'N', key=True)
    kv(ws, 2, 15, 'μ with bias', '=IF(UPPER($L$14)="Y",$H$9+$L$13,$H$9)', '0.000')

    kv(ws, 3, 4, 'Side held', None, inp=True)
    kv(ws, 3, 5, 'Contracts', None, NUM0, inp=True)
    kv(ws, 3, 6, 'Avg entry ¢', None, CENT, inp=True)
    kv(ws, 3, 7, 'Cost $', '=IF($Q$5="","",$Q$5*$Q$6/100)', MONEY)
    kv(ws, 3, 8, 'Fees paid $', None, MONEY, inp=True)
    kv(ws, 3, 9, 'Settled', None, inp=True)
    kv(ws, 3, 10, 'Payout $',
       '=IF(OR($Q$9="",$Q$4=""),"",IF(UPPER($Q$9)=UPPER($Q$4),$Q$5,0))', MONEY)
    kv(ws, 3, 11, 'Realised P&L $', '=IF($Q$10="","",$Q$10-$Q$7-IF($Q$8="",0,$Q$8))',
       MONEY, bold=True)
    kv(ws, 3, 12, 'ROI %', '=IF(OR($Q$11="",$Q$7=""),"",$Q$11/$Q$7)', PCT, bold=True)
    kv(ws, 3, 13, 'Risk cap $ / market', 300, MONEY, key=True)
    kv(ws, 3, 14, 'Book depth at touch $', None, MONEY, inp=True)
    kv(ws, 3, 15, 'Exposure vs cap', '=IF($Q$7="","",$Q$7/$Q$13)', PCT)

    heads = ['Date', 'Actual °C', 'Expectation', 'Dev', 'Dev3', 'MTD mean',
             str(d['tops'][0]), str(d['tops'][1]), 'Needed', 'P(YES)', 'Fair NO',
             'Mkt YES', 'Edge', 'Volume', 'Open int', 'Signal', 'Note',
             'Pace', 'Model ¢']
    for L, h in zip('ABCDEFGHIJKLMNOPQRS', heads):
        c = ws[f'{L}{HDR}']
        c.value, c.font = h, F(size=9, bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[HDR].height = 26

    r_ = PRE0
    for row in d['rows']:
        inm = row['day'] is not None
        ws[f'A{r_}'] = row['label']
        ws[f'B{r_}'] = row['actual']
        ws[f'C{r_}'] = row['fcst']
        ws[f'D{r_}'] = f'=IF(OR($B{r_}="",$C{r_}=""),"",$B{r_}-$C{r_})'
        ws[f'E{r_}'] = f'=IF(COUNT($D{r_-2}:$D{r_})<3,"",AVERAGE($D{r_-2}:$D{r_}))'
        ws[f'G{r_}'], ws[f'H{r_}'] = row['an'][0], row['an'][1]
        if inm:
            dn = row['day']
            ws[f'F{r_}'] = f'=IF($B{r_}="","",AVERAGE($B${M0}:$B{r_}))'
            ws[f'I{r_}'] = (f'=IF(OR($B{r_}="",{dn}>={nd}),"",'
                            f'($D$9*{nd}-SUM($B${M0}:$B{r_}))/({nd}-{dn}))')
            ws[f'J{r_}'] = (f'=IF($B{r_}="","",1-NORMDIST($D$5,$D$6+$D$7*'
                            f'((SUM($B${M0}:$B{r_})+({nd}-{dn})*$H$7)/{nd}),$D$11,TRUE))')
            ws[f'R{r_}'] = '=$D$9'
        ws[f'K{r_}'] = f'=IF($J{r_}="","",100*(1-$J{r_}))'
        ws[f'L{r_}'] = row['mkt'].get('close')
        ws[f'M{r_}'] = f'=IF(OR($J{r_}="",$L{r_}=""),"",$L{r_}-100*$J{r_})'
        ws[f'N{r_}'] = row['mkt'].get('vol')
        ws[f'O{r_}'] = row['mkt'].get('oi')
        ws[f'P{r_}'] = (f'=IF($E{r_}="","",IF($E{r_}>0.05,"WARM dev3 — spike risk UP",'
                        f'IF($E{r_}<-0.05,"COLD dev3 — spike risk DOWN","")))')
        ws[f'S{r_}'] = f'=IF($J{r_}="","",$J{r_}*100)'
        for L in 'ABCDEFGHIJKLMNOPQRS':
            c = ws[f'{L}{r_}']
            c.border, c.font = BOX, F(size=9)
            if not inm:
                c.fill = PatternFill('solid', fgColor=GREY)
        for L in ('B', 'C', 'L', 'N', 'O'):
            ws[f'{L}{r_}'].font = F(size=9, color=BLUE_TXT)
            if inm:
                ws[f'{L}{r_}'].fill = PatternFill('solid', fgColor=GREEN)
        for L in 'BCDEFGHIR':
            ws[f'{L}{r_}'].number_format = ANOM
        for L in ('K', 'L', 'M', 'S'):
            ws[f'{L}{r_}'].number_format = CENT
        ws[f'J{r_}'].number_format = PCT
        for L in ('N', 'O'):
            ws[f'{L}{r_}'].number_format = NUM0
        ws[f'A{r_}'].alignment = Alignment(horizontal='left')
        ws[f'P{r_}'].font = F(size=8, italic=True, color=RED)
        r_ += 1

    # MEAN row: B-K average the in-month days only; L-O include the pre-month
    # rows because the market already trades there.
    mr = M1 + 1
    a_ = ws.cell(mr, 1, 'MEAN')
    a_.font = F(size=9, bold=True)
    a_.fill = PatternFill('solid', fgColor=BAND)
    a_.border = BOX
    for col, first, fmt in (('B', M0, ANOM), ('C', M0, ANOM), ('D', M0, ANOM),
                            ('E', M0, ANOM), ('F', M0, ANOM), ('G', M0, ANOM),
                            ('H', M0, ANOM), ('J', M0, '0%'), ('K', M0, '0.00'),
                            ('L', PRE0, '0.00'), ('M', PRE0, '0.00'),
                            ('N', PRE0, '0.00'), ('O', PRE0, '0.00')):
        c_ = ws[f'{col}{mr}']
        c_.value = f'=IFERROR(AVERAGE({col}{first}:{col}{M1}),"")'
        c_.number_format = fmt
        c_.font = F(size=9)
        c_.fill = PatternFill('solid', fgColor=BAND)
        c_.border = BOX
    for col in ('I', 'P', 'Q', 'R', 'S'):        # averaging these means nothing
        c_ = ws[f'{col}{mr}']
        c_.fill = PatternFill('solid', fgColor=BAND)
        c_.border = BOX

    for rng, op, val, col in ((f'E{PRE0}:E{M1}', 'greaterThan', '0.05', 'FFF8CBAD'),
                              (f'E{PRE0}:E{M1}', 'lessThan', '-0.05', 'FFBDD7EE'),
                              (f'M{PRE0}:M{M1}', 'greaterThanOrEqual', '3', 'FFC6EFCE'),
                              (f'M{PRE0}:M{M1}', 'lessThanOrEqual', '-3', 'FFFFC7CE')):
        ws.conditional_formatting.add(rng, CellIsRule(
            operator=op, formula=[val], fill=PatternFill('solid', bgColor=col)))
    ws.conditional_formatting.add(f'B{M0}:B{M1}', CellIsRule(
        operator='greaterThanOrEqual', formula=['$D$9'],
        fill=PatternFill('solid', bgColor='FFC6EFCE')))
    ws.conditional_formatting.add('D12', CellIsRule(
        operator='equal', formula=['"DRIFT"'],
        fill=PatternFill('solid', bgColor='FFFFC7CE'), font=Font(bold=True)))
    exp_cf(ws, d['nd'])

    banner(ws, f'A{CAL}:C{CAL}', f"CALIBRATION — {d['mname']} 1990–{d['year']-1}", 9)
    for i, h in enumerate(['Year', 'ERA5 mean', 'NOAA']):
        c = ws.cell(CAL + 1, 1 + i, h)
        c.font, c.fill = F(size=9, bold=True, color=WHITE), PatternFill('solid', fgColor=NAVY)
    for j, (yy, e5, nn) in enumerate(d['hist']):
        ws.cell(CAL + 2 + j, 1, yy).number_format = '0'
        ws.cell(CAL + 2 + j, 2, e5).number_format = ANOM
        ws.cell(CAL + 2 + j, 3, nn).number_format = '0.00'

    banner(ws, f'E{CAL}:F{CAL}', 'σ_f BY k', 9)
    for i, h in enumerate(['k', 'σ_f']):
        c = ws.cell(CAL + 1, 5 + i, h)
        c.font, c.fill = F(size=9, bold=True, color=WHITE), PatternFill('solid', fgColor=NAVY)
    for j, kk in enumerate(sorted(d['coll'])):
        ws.cell(CAL + 2 + j, 5, kk).number_format = '0'
        ws.cell(CAL + 2 + j, 6, d['coll'][kk]).number_format = '0.0000'

    banner(ws, f'H{CAL}:L{CAL}', 'TRANSLATION BIAS — 2026 prints', 9)
    for i, h in enumerate(['Month', 'ERA5', 'NOAA', 'Fitted', 'Residual']):
        c = ws.cell(CAL + 1, 8 + i, h)
        c.font, c.fill = F(size=9, bold=True, color=WHITE), PatternFill('solid', fgColor=NAVY)
    for j, b_ in enumerate(d['bias']):
        for i, v in enumerate(b_):
            ws.cell(CAL + 2 + j, 8 + i, v)
    bn = CAL + 2 + len(d['bias'])
    ws.cell(bn, 8, 'MEAN').font = F(size=9, bold=True)
    ws.cell(bn, 12, f'=AVERAGE(L{CAL+2}:L{bn-1})').number_format = '+0.0000;-0.0000'
    return ws


# ------------------------------------------------------------------ charts
def ensure_helpers(ws, d):
    """Columns R (record pace) and S (model YES ¢) exist only to give the charts
    clean series. Sheets built before these existed get them added here."""
    M1 = geom(d['nd'])['m1']
    if ws[f'R{HDR}'].value == 'Pace':
        return 0
    added = 0
    for L, h, w in (('R', 'Pace', 6.5), ('S', 'Model ¢', 8.0)):
        c = ws[f'{L}{HDR}']
        c.value, c.font = h, F(size=9, bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = BOX
        ws.column_dimensions[L].width = w
    for r in range(PRE0, M1 + 1):
        if r >= M0:
            ws[f'R{r}'] = '=$D$9'
            ws[f'R{r}'].number_format = ANOM
        ws[f'S{r}'] = f'=IF($J{r}="","",$J{r}*100)'
        ws[f'S{r}'].number_format = CENT
        for L in ('R', 'S'):
            ws[f'{L}{r}'].font = F(size=9)
            ws[f'{L}{r}'].border = BOX
        added += 1
    return added


def rebuild_charts(ws, d):
    g = geom(d['nd'])
    M1, CAL = g['m1'], g['cal']
    ensure_helpers(ws, d)
    ws._charts = []

    def style(ch, skip):
        ch.height, ch.width = 8.5, 18
        ch.x_axis.delete = ch.y_axis.delete = False
        ch.y_axis.majorGridlines = ch.x_axis.majorGridlines = None
        ch.x_axis.tickLblSkip = skip
        ch.x_axis.tickMarkSkip = skip
        ch.legend.position = 'b'
        ch.legend.overlay = False

    def ser(col, r0, r1, title):
        s = Series(Reference(ws, min_col=col, min_row=r0, max_row=r1), title=title)
        s.smooth = False
        return s

    cats_m = Reference(ws, min_col=1, min_row=M0, max_row=M1)
    cats_a = Reference(ws, min_col=1, min_row=PRE0, max_row=M1)

    c1 = LineChart()
    c1.title = f"Month-to-date vs record pace — {d['mname']} {d['year']} (°C vs 1991–2020)"
    for col, ttl in ((6, 'MTD mean'), (7, str(d['tops'][0])), (8, str(d['tops'][1])),
                     (18, f"record pace {d['bar']:+.3f}")):
        c1.series.append(ser(col, M0, M1, ttl))
    c1.set_categories(cats_m)
    style(c1, 3)
    ws.add_chart(c1, f'A{CAL+40}')

    c2 = LineChart()
    c2.title = "Model vs market — YES in cents (gap = edge)"
    c2.series.append(ser(19, PRE0, M1, 'model YES ¢'))
    c2.series.append(ser(12, PRE0, M1, 'market YES ¢'))
    c2.set_categories(cats_a)
    style(c2, 3)
    ws.add_chart(c2, f'K{CAL+40}')

    c3 = BarChart()
    c3.type = 'col'
    c3.title = f"Volume and open interest — {d['ticker']}"
    c3.series.append(ser(14, PRE0, M1, 'volume'))
    c3.set_categories(cats_a)
    style(c3, 3)
    l3 = LineChart()
    l3.series.append(ser(15, PRE0, M1, 'open interest'))
    c3 += l3
    ws.add_chart(c3, f'A{CAL+58}')

    c4 = LineChart()
    c4.title = "Forecast error — Dev and Dev3 (spike tripwire, ±0.05)"
    c4.series.append(ser(4, PRE0, M1, 'Dev'))
    c4.series.append(ser(5, PRE0, M1, 'Dev3'))
    c4.set_categories(cats_a)
    style(c4, 3)
    ws.add_chart(c4, f'K{CAL+58}')


# ------------------------------------------------------------------ refresh
def refresh(ws, d):
    """Write ONLY input cells. Never a formula, never formatting, never a chart."""
    g = geom(d['nd'])
    touched = 0
    r_ = PRE0
    for row in d['rows']:
        for col, val in (('B', row['actual']), ('C', row['fcst']),
                         ('L', row['mkt'].get('close')), ('N', row['mkt'].get('vol')),
                         ('O', row['mkt'].get('oi')),
                         ('G', row['an'][0]), ('H', row['an'][1])):
            if val is not None and ws[f'{col}{r_}'].value != val:
                ws[f'{col}{r_}'] = val
                touched += 1
        r_ += 1
    e = d['ens']
    if e:
        for cell, val in (('L4', e['run']), ('L5', e['n']), ('L6', e['cold']),
                          ('L7', e['avg']), ('L8', e['hot']), ('L9', e['breach']),
                          ('L10', e['p'])):
            if ws[cell].value != val:
                ws[cell] = val
                touched += 1
    if d['pipe_p'] is not None and ws['H11'].value != d['pipe_p']:
        ws['H11'] = d['pipe_p']
        touched += 1
    return touched


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=BOOK)
    ap.add_argument('--month')
    ap.add_argument('--build', action='store_true')
    ap.add_argument('--auto', action='store_true',
                    help='build the sheet if missing, then refresh (for the daily job)')
    ap.add_argument('--charts', action='store_true')
    a = ap.parse_args()

    if a.month:
        yr, mo = map(int, a.month.split('-'))
    else:
        now = dt.datetime.now()
        yr, mo = now.year, now.month

    lock = os.path.join(os.path.dirname(a.file), '~$' + os.path.basename(a.file))
    if os.path.exists(lock):
        sys.exit(f"REFUSING TO WRITE: {os.path.basename(a.file)} is open in Excel "
                 f"({os.path.basename(lock)} present). Close it and re-run.")

    d = collect(yr, mo)
    wb = openpyxl.load_workbook(a.file)
    name = f'{calendar.month_abbr[mo]}_{yr}'

    if a.auto and name not in wb.sheetnames:
        a.build = True
    if a.build:
        if name in wb.sheetnames:
            sys.exit(f"{name} already exists — refusing to overwrite. "
                     f"Delete the tab first if you really want it rebuilt.")
        ws = build_sheet(wb, d)
        rebuild_charts(ws, d)
        print(f'built {name}: {d["nd"]} days, rows {M0}–{geom(d["nd"])["m1"]}, '
              f'{len(d["hist"])} calibration years')
        refresh(ws, d)
    else:
        if name not in wb.sheetnames:
            sys.exit(f"{name} does not exist — run with --build first.")
        ws = wb[name]
        n = refresh(ws, d)
        if exp_cf(ws, d['nd']):        # one-time: only this month's sheet
            print(f'  {name}: Expectation under/over colour rule added')
        print(f'{name}: {n} cell(s) updated '
              f'(k={sum(1 for r in d["rows"] if r["day"] and r["actual"] is not None)}, '
              f'ens={"yes" if d["ens"] else "none"}, '
              f'pipeline P={d["pipe_p"]})')
        if a.charts:
            rebuild_charts(ws, d)
            print('  charts rebuilt')

    wb.save(a.file)
    print('saved', os.path.basename(a.file))


if __name__ == '__main__':
    main()
